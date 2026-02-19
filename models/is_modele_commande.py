# -*- coding: utf-8 -*-
from odoo import api, fields, models, _                                      # type: ignore
from odoo.exceptions import UserError, ValidationError, Warning              # type: ignore
from odoo.tools.float_utils import float_compare, float_is_zero, float_round # type: ignore
from odoo.osv import expression                                              # type: ignore
import time
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Font, Color, Fill, Alignment,PatternFill
from copy import copy
import base64
import datetime
import logging
_logger = logging.getLogger(__name__)


class IsModeleCommandeLigne(models.Model):
    _name = 'is.modele.commande.ligne'
    _description = "Lignes des modèles de commandes"
    _order='modele_id,sequence,product_id'


    @api.depends('product_id')
    def _compute(self):
        for obj in self:
            weight = 0
            if obj.product_id:
                weight = obj.product_id.is_poids_net_colis / (obj.product_id.is_nb_pieces_par_colis or 1)
            obj.weight          = weight
            obj.product_name    = obj.product_id.name
            obj.default_code    = obj.product_id.default_code
            obj.ref_fournisseur = obj.product_id.is_ref_fournisseur
            #** Limite fournisseur ********************************************
            filtre=[
                ('product_tmpl_id', '=', obj.product_id.product_tmpl_id.id),
            ]
            heure_envoi_id=False
            suppliers=self.env['product.supplierinfo'].search(filtre,limit=1)
            for s in suppliers:
                heure_envoi_id = s.name.is_heure_envoi_id.id
            obj.heure_envoi_id = heure_envoi_id
            #******************************************************************


    @api.depends('product_id')
    def compute_price_unit(self):
        company = self.env.user.company_id
        for obj in self:
            prix=False
            if obj.modele_id.prix_futur and obj.product_id:
                name = "is_prix_vente_futur_%s"%obj.modele_id.prix_futur
                prix = getattr(obj.product_id, name)
            else:
                partner = obj.modele_id.partner_id
                if partner and obj.product_id:
                    pricelist=partner.property_product_pricelist
                    product = obj.product_id.with_context(
                        partner=partner,
                        quantity=1,
                        date=datetime.date.today(),
                        pricelist=pricelist,
                        uom=obj.product_id.uom_id.id
                    )
                    product_context = dict(
                        self.env.context, 
                        partner_id=partner.id, 
                        date=datetime.date.today(), 
                        uom=obj.product_id.uom_id.id
                    )
                    prix, rule_id = pricelist.with_context(product_context).get_product_price_rule(product, 1.0, partner)

            print(obj.modele_id.name,obj.default_code, obj.modele_id.prix_futur,prix)

            obj.price_unit = prix


    modele_id       = fields.Many2one('is.modele.commande', 'Modèle de commandes', required=True, ondelete='cascade')
    sequence        = fields.Integer('Séquence')
    product_id      = fields.Many2one('product.product', 'Article', required=True, index=True)
    product_name    = fields.Char('Désignation article'                    , compute='_compute', readonly=True, store=True)
    default_code    = fields.Char('Réf Fromtome'                           , compute='_compute', readonly=True, store=True)
    ref_fournisseur = fields.Char('Réf Fournisseur'                        , compute='_compute', readonly=True, store=True)
    heure_envoi_id  = fields.Many2one('is.heure.maxi', 'Limite fournisseur', compute='_compute', readonly=True, store=True)
    weight          = fields.Float(string='Poids unitaire', digits='Stock Weight', compute='_compute', readonly=True, store=True)
    qt_livree       = fields.Float(string='Qt livrée', help="quantité livrée au moment de l'initialisation", readonly=True)
    price_unit      = fields.Float("Prix", digits='Product Price', compute='compute_price_unit', readonly=True, store=True)
    alerte          = fields.Boolean('Alerte', compute='_compute_alerte')
    is_mise_en_avant = fields.Boolean(related="product_id.is_mise_en_avant")
    is_preco         = fields.Boolean(related="product_id.is_preco")


    @api.depends('product_id')
    def _compute_alerte(self):
        for obj in self:
            alerte=False
            if obj.product_id.active==False:
                alerte=True
            obj.alerte=alerte
         

    def alerte_action(self):
        for obj in self:
             print(obj)


class IsModeleCommande(models.Model):
    _name        = 'is.modele.commande'
    _description = "Modèle de commandes"
    _order       ='name'

    name                = fields.Char('Nom du modèle', required=True)
    partner_id          = fields.Many2one('res.partner', 'Client')
    enseigne_id         = fields.Many2one(related='partner_id.is_enseigne_id')
    modele_commande_ids = fields.Many2many('ir.attachment', 'is_modele_commande_modele_commande_rel', 'enseigne_id', 'file_id', 'Modèle de commande client')
    ligne_ids           = fields.One2many('is.modele.commande.ligne', 'modele_id', 'Lignes')
    alerte              = fields.Boolean('Alerte', compute='_compute_alerte')
    product_ids         = fields.Many2many('product.product', 'is_modele_commande_product_rel', 'modele_id', 'product_id', 'Articles à ajouter')

    prix_futur = fields.Selection([
            ('ft'        , 'FROMTOME'),
            ('cdf_quai'  , 'A QUAI'),
            ('cdf_franco', 'FRANCO'),
        ], 'Prix futur à afficher', copy=False)
    prix_futur_a_partir_du = fields.Date('Prix futur à partir du', copy=False)


    def write(self, vals):
        if 'prix_futur' in vals and not vals['prix_futur']:
            vals['prix_futur_a_partir_du'] = False
        return super().write(vals)


    def ajouter_articles_action(self):
        for obj in self:
            product_ids=[]
            for l in obj.ligne_ids:
                if l.product_id not in product_ids:
                    product_ids.append(l.product_id)
            new_product_ids=[]
            for p in obj.product_ids:
                if p not in product_ids:
                    new_product_ids.append(p._origin)
            for p in new_product_ids:
                vals={
                    'modele_id'   : obj._origin.id,
                    'product_id' : p.id,
                }
                self.env['is.modele.commande.ligne'].create(vals)
            obj.product_ids=False


    @api.depends('ligne_ids')
    def _compute_alerte(self):
        for obj in self:
            alerte=False
            for line in obj.ligne_ids:
                if line.product_id.active==False:
                    alerte=True
                    break
            obj.alerte=alerte
         

    def alerte_action(self):
        for obj in self:
             print(obj)


    def actualiser_modele_excel_action(self):
        for obj in self:
            if not obj.enseigne_id:
                raise Warning("Enseigne obligatoire pour générer la commande Excel !")
            for line in obj.ligne_ids:
                line._compute()
            for modele in obj.enseigne_id.modele_commande_ids:
                name = obj.name

                #** Enregistrement du modèle en local *************************
                res = modele.datas
                res = base64.b64decode(res)
                path="/tmp/modele-commande-%s.xlsx"%obj.id
                f = open(path,'wb')
                f.write(res)
                f.close()
                #**************************************************************

                #** Ecritures des lignes **************************************
                wb = load_workbook(path)
                sheet = wb.active

                #** Affichage Prix ou Prix à partir du ************************
                val = "Prix"
                if obj.prix_futur and obj.prix_futur_a_partir_du:
                    val = "Prix à partir du %s"%(obj.prix_futur_a_partir_du.strftime("%d/%m/%y"))
                sheet.cell(row=6, column=5).value = val
                #**************************************************************



                row=7 #Première ligne pour enregistrer les données
                lig=1
                filtre=[
                    ('modele_id','=',obj.id),
                ]
                lines = self.env['is.modele.commande.ligne'].search(filtre, order="product_name")
                for line in lines:
                    if line.product_id.default_code:

                        line.compute_price_unit()


                        sheet.cell(row=row, column=1).value = lig
                        sheet.cell(row=row, column=2).value = line.product_id.name
                        sheet.cell(row=row, column=3).value = line.product_id.default_code
                        sheet.cell(row=row, column=4).value = line.product_id.is_ref_fournisseur or ''
                        sheet.cell(row=row, column=5).value = line.price_unit or ''
                        sheet.cell(row=row, column=6).value = line.product_id.uom_id.name
                        sheet.cell(row=row, column=8).value = line.heure_envoi_id.name or ''

                        sheet.cell(row=row, column=10).value = line.product_id.is_nb_pieces_par_colis
                        sheet.cell(row=row, column=11).value = line.product_id.is_poids_net_colis

                        #** Ajout de la formule pour le montant ***************
                        l=lig+6
                        key_montant = sheet.cell(row=row, column=9).coordinate
                        formule='=IF(F%s="Pc",E%s*G%s*J%s,E%s*K%s*G%s)'%(l,l,l,l,l,l,l)
                        sheet[key_montant]=formule
                        #******************************************************

                        if line.is_mise_en_avant or line.is_preco:
                            if line.is_mise_en_avant:
                                txt = "Produit mis en avant"
                            if line.is_preco:
                                txt = "Produit recommandé"
                            sheet.cell(row=row, column=1).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=2).alignment = Alignment(vertical='center', horizontal='left') 
                            sheet.cell(row=row, column=3).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=4).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=5).alignment = Alignment(vertical='center', horizontal='right') 
                            sheet.cell(row=row, column=6).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=2).value = "%s:\n%s"%(txt,line.product_id.name)
                            sheet.row_dimensions[row].height=34 
                            for i in  range(1,10):
                                cell = sheet.cell(row=row,column=i)
                                font = copy(cell.font)
                                font.size = 12
                                cell.font = font
                                fill = PatternFill(fill_type='solid', start_color='FFFBCC', end_color='FFFBCC')
                                cell.fill=fill
                        row+=1
                        lig+=1

                sheet.print_area = 'A1:I1000'
                wb.save(path)
                #**************************************************************

                # ** Creation ou modification de la pièce jointe **************
                attachments = obj.modele_commande_ids
                name="%s.xlsx"%(obj.name)
                model=self._name
                file = open(path,'rb').read()
                datas = base64.b64encode(file)
                vals = {
                    'name':        name,
                    'type':        'binary',
                    'res_model':   model,
                    'res_id':      obj.id,
                    'datas':       datas,
                }
                if len(attachments):
                    attachment=attachments[0]
                    attachment.write(vals)
                else:
                    attachment = self.env['ir.attachment'].create(vals)
                obj.modele_commande_ids=[attachment.id]
                #**************************************************************


    def initialiser_action(self):
        cr = self._cr
        for obj in self:
            filtre=[
                ('is_modele_commande_id','=',obj.id),
            ]
            partners = self.env['res.partner'].search(filtre)
            if len(partners)>0:
                obj.ligne_ids.unlink()
                ids=[]
                for partner in partners:
                    ids.append("'%s'"%(partner.id))
                ids=','.join(ids)
                sql="""
                    SELECT  
                        sol.product_id,
                        sum(sol.qty_delivered) qt_livree
                    FROM sale_order so join sale_order_line sol on so.id=sol.order_id
                                    join product_product pp on sol.product_id=pp.id
                    WHERE 
                        so.state='sale' and so.partner_id in ("""+ids+""") and sol.qty_delivered>0
                    GROUP BY sol.product_id 
                """
                cr.execute(sql)
                for row in cr.dictfetchall():
                    vals={
                        'modele_id' : obj.id,
                        'product_id': row["product_id"],
                        'qt_livree' : row["qt_livree"],
                    }
                    self.env['is.modele.commande.ligne'].create(vals)
                obj.trier_action()


    def trier_action(self):
        for obj in self:
            for line in obj.ligne_ids:
                line.sequence = - line.weight*10000
