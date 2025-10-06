# -*- coding: utf-8 -*-
from odoo import api, fields, models, _                                      # type: ignore
from odoo.exceptions import UserError, ValidationError, Warning              # type: ignore
from odoo.tools.float_utils import float_compare, float_is_zero, float_round # type: ignore
from odoo.osv import expression                                              # type: ignore
import time
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Font, Color, Fill, Alignment,PatternFill
from copy import copy
import base64
import datetime
import logging
_logger = logging.getLogger(__name__)


class IsModeleCommandeLigne(models.Model):
    _name = 'is.modele.commande.ligne'
    _description = "Lignes des modèles de commandes"
    _order='modele_id,sequence,product_id'


    @api.depends('product_id')
    def _compute(self):
        for obj in self:
            weight = 0
            if obj.product_id:
                weight = obj.product_id.is_poids_net_colis / (obj.product_id.is_nb_pieces_par_colis or 1)
            obj.weight          = weight
            obj.product_name    = obj.product_id.name
            obj.default_code    = obj.product_id.default_code
            obj.ref_fournisseur = obj.product_id.is_ref_fournisseur
            #** Limite fournisseur ********************************************
            filtre=[
                ('product_tmpl_id', '=', obj.product_id.product_tmpl_id.id),
            ]
            heure_envoi_id=False
            suppliers=self.env['product.supplierinfo'].search(filtre,limit=1)
            for s in suppliers:
                heure_envoi_id = s.name.is_heure_envoi_id.id
            obj.heure_envoi_id = heure_envoi_id
            #******************************************************************


    @api.depends('product_id')
    def _compute_price_unit(self):
        company = self.env.user.company_id
        for obj in self:
            prix=False
            partner = obj.modele_id.partner_id
            if partner and obj.product_id:
                pricelist=partner.property_product_pricelist
                product = obj.product_id.with_context(
                    partner=partner,
                    quantity=1,
                    date=datetime.date.today(),
                    pricelist=pricelist,
                    uom=obj.product_id.uom_id.id
                )
                product_context = dict(
                    self.env.context, 
                    partner_id=partner.id, 
                    date=datetime.date.today(), 
                    uom=obj.product_id.uom_id.id
                )
                prix, rule_id = pricelist.with_context(product_context).get_product_price_rule(product, 1.0, partner)
            obj.price_unit = prix


    modele_id       = fields.Many2one('is.modele.commande', 'Modèle de commandes', required=True, ondelete='cascade')
    sequence        = fields.Integer('Séquence')
    product_id      = fields.Many2one('product.product', 'Article', required=True)
    product_name    = fields.Char('Désignation article'                    , compute='_compute', readonly=True, store=True)
    default_code    = fields.Char('Réf Fromtome'                           , compute='_compute', readonly=True, store=True)
    ref_fournisseur = fields.Char('Réf Fournisseur'                        , compute='_compute', readonly=True, store=True)
    heure_envoi_id  = fields.Many2one('is.heure.maxi', 'Limite fournisseur', compute='_compute', readonly=True, store=True)
    weight          = fields.Float(string='Poids unitaire', digits='Stock Weight', compute='_compute', readonly=True, store=True)
    qt_livree       = fields.Float(string='Qt livrée', help="quantité livrée au moment de l'initialisation", readonly=True)
    price_unit      = fields.Float("Prix", digits='Product Price', compute='_compute_price_unit', readonly=True, store=False)
    alerte          = fields.Boolean('Alerte', compute='_compute_alerte')
    is_mise_en_avant = fields.Boolean(related="product_id.is_mise_en_avant")
    is_preco         = fields.Boolean(related="product_id.is_preco")


    @api.depends('product_id')
    def _compute_alerte(self):
        for obj in self:
            alerte=False
            if obj.product_id.active==False:
                alerte=True
            obj.alerte=alerte
         

    def alerte_action(self):
        for obj in self:
             print(obj)


class IsModeleCommande(models.Model):
    _name        = 'is.modele.commande'
    _description = "Modèle de commandes"
    _order       ='name'

    name                = fields.Char('Nom du modèle', required=True)
    partner_id          = fields.Many2one('res.partner', 'Client')
    enseigne_id         = fields.Many2one(related='partner_id.is_enseigne_id')
    modele_commande_ids = fields.Many2many('ir.attachment', 'is_modele_commande_modele_commande_rel', 'enseigne_id', 'file_id', 'Modèle de commande client')
    ligne_ids           = fields.One2many('is.modele.commande.ligne', 'modele_id', 'Lignes')
    alerte              = fields.Boolean('Alerte', compute='_compute_alerte')


    @api.depends('ligne_ids')
    def _compute_alerte(self):
        for obj in self:
            alerte=False
            for line in obj.ligne_ids:
                if line.product_id.active==False:
                    alerte=True
                    break
            obj.alerte=alerte
         

    def alerte_action(self):
        for obj in self:
             print(obj)


    def actualiser_modele_excel_action(self):
        for obj in self:
            if not obj.enseigne_id:
                raise Warning("Enseigne obligatoire pour générer la commande Excel !")
            for line in obj.ligne_ids:
                line._compute()
            for modele in obj.enseigne_id.modele_commande_ids:
                name = obj.name

                #** Enregistrement du modèle en local *************************
                res = modele.datas
                res = base64.b64decode(res)
                path="/tmp/modele-commande-%s.xlsx"%obj.id
                f = open(path,'wb')
                f.write(res)
                f.close()
                #**************************************************************

                #** Ecritures des lignes **************************************
                wb = load_workbook(path)
                sheet = wb.active
                row=7 #Première ligne pour enregistrer les données
                lig=1
                filtre=[
                    ('modele_id','=',obj.id),
                ]
                lines = self.env['is.modele.commande.ligne'].search(filtre, order="product_name")
                for line in lines:
                    if line.product_id.default_code:
                        sheet.cell(row=row, column=1).value = lig
                        sheet.cell(row=row, column=2).value = line.product_id.name
                        sheet.cell(row=row, column=3).value = line.product_id.default_code
                        sheet.cell(row=row, column=4).value = line.product_id.is_ref_fournisseur or ''
                        sheet.cell(row=row, column=5).value = line.price_unit or ''
                        sheet.cell(row=row, column=6).value = line.product_id.uom_id.name
                        sheet.cell(row=row, column=8).value = line.heure_envoi_id.name or ''

                        sheet.cell(row=row, column=10).value = line.product_id.is_nb_pieces_par_colis
                        sheet.cell(row=row, column=11).value = line.product_id.is_poids_net_colis




                        #** Ajout de la formule pour le montant ***************

                        #formule='SI(F8="Pc";E8*G8*J8;E8*K8*G8)'
                        l=lig+6
                        key_montant = sheet.cell(row=row, column=9).coordinate
                        formule='=IF(F%s="Pc",E%s*G%s*J%s,E%s*K%s*G%s)'%(l,l,l,l,l,l,l)
                        sheet[key_montant]=formule


                        # key_price   = sheet.cell(row=row, column=5).coordinate
                        # key_qty     = sheet.cell(row=row, column=7).coordinate
                        # key_montant = sheet.cell(row=row, column=9).coordinate
                        # formule = "=%s*%s"%(key_price,key_qty)
                        # sheet[key_montant]=formule
                        #******************************************************

                        if line.is_mise_en_avant or line.is_preco:
                            if line.is_mise_en_avant:
                                txt = "Produit mis en avant"
                            if line.is_preco:
                                txt = "Produit recommandé"
                            sheet.cell(row=row, column=1).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=2).alignment = Alignment(vertical='center', horizontal='left') 
                            sheet.cell(row=row, column=3).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=4).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=5).alignment = Alignment(vertical='center', horizontal='right') 
                            sheet.cell(row=row, column=6).alignment = Alignment(vertical='center', horizontal='center') 
                            sheet.cell(row=row, column=2).value = "%s:\n%s"%(txt,line.product_id.name)
                            sheet.row_dimensions[row].height=34 
                            for i in  range(1,10):
                                cell = sheet.cell(row=row,column=i)
                                font = copy(cell.font)
                                font.size = 12
                                cell.font = font
                                fill = PatternFill(fill_type='solid', start_color='FFFBCC', end_color='FFFBCC')
                                cell.fill=fill
                        row+=1
                        lig+=1

                sheet.print_area = 'A1:I1000'
                wb.save(path)
                #**************************************************************

                # ** Creation ou modification de la pièce jointe **************
                attachments = obj.modele_commande_ids
                name="%s.xlsx"%(obj.name)
                model=self._name
                file = open(path,'rb').read()
                datas = base64.b64encode(file)
                vals = {
                    'name':        name,
                    'type':        'binary',
                    'res_model':   model,
                    'res_id':      obj.id,
                    'datas':       datas,
                }
                if len(attachments):
                    attachment=attachments[0]
                    attachment.write(vals)
                else:
                    attachment = self.env['ir.attachment'].create(vals)
                obj.modele_commande_ids=[attachment.id]
                #**************************************************************


    def initialiser_action(self):
        cr = self._cr
        for obj in self:
            filtre=[
                ('is_modele_commande_id','=',obj.id),
            ]
            partners = self.env['res.partner'].search(filtre)
            if len(partners)>0:
                obj.ligne_ids.unlink()
                ids=[]
                for partner in partners:
                    ids.append("'%s'"%(partner.id))
                ids=','.join(ids)
                sql="""
                    SELECT  
                        sol.product_id,
                        sum(sol.qty_delivered) qt_livree
                    FROM sale_order so join sale_order_line sol on so.id=sol.order_id
                                    join product_product pp on sol.product_id=pp.id
                    WHERE 
                        so.state='sale' and so.partner_id in ("""+ids+""") and sol.qty_delivered>0
                    GROUP BY sol.product_id 
                """
                cr.execute(sql)
                for row in cr.dictfetchall():
                    vals={
                        'modele_id' : obj.id,
                        'product_id': row["product_id"],
                        'qt_livree' : row["qt_livree"],
                    }
                    self.env['is.modele.commande.ligne'].create(vals)
                obj.trier_action()


    def trier_action(self):
        for obj in self:
            for line in obj.ligne_ids:
                line.sequence = - line.weight*10000


class SaleOrderLine(models.Model):
    _inherit = "sale.order.line"

    def get_nb_colis(self):
        nb        = self.product_id.is_nb_pieces_par_colis
        poids_net = self.product_id.is_poids_net_colis
        unite     = self.product_uom.category_id.name
        nb_colis  = 0
        if unite=="Poids":
            if poids_net>0:
                nb_colis = self.product_uom_qty/poids_net
        else:
            if nb>0:
                nb_colis = self.product_uom_qty / nb
        return round(nb_colis)


    def get_discount(self):
        for obj in self:
            now = obj.order_id.is_date_livraison or datetime.date.today()
            discount = 0

            #** Recherche promo client ****************************************
            filtre=[
                    ('date_debut_promo', '<=', now),
                    ('date_fin_promo', '>=', now),
                    ('partner_id', '=', obj.order_id.partner_id.id),
                    ('promo_generique','=',False),
                ]
            promos=self.env['is.promo.client'].search(filtre, limit=1)
            if len(promos)==0:
                #** Recherche promo générique *********************************
                filtre=[
                        ('date_debut_promo', '<=', now),
                        ('date_fin_promo', '>=', now),
                        ('pricelist_id', '=', obj.order_id.pricelist_id.id),
                        ('promo_generique','=',True),
                    ]
                promos=self.env['is.promo.client'].search(filtre, limit=1)
            for promo in promos:
                filtre=[
                        ('promo_id'        , '=', promo.id),
                        ('product_id'      , '=', obj.product_id.id),
                        ('date_debut_promo', '<=', now),
                        ('date_fin_promo'  , '>=', now),
                    ]
                lignes=self.env['is.promo.client.ligne'].search(filtre, limit=1)
                for ligne in lignes:
                    discount = ligne.remise_client
            #******************************************************************


            #** Recherche remises particulières *******************************
            filtre=[
                    ('partner_id', '=', obj.order_id.partner_id.id),
                    ('product_id', '=', obj.product_id.id),
                ]
            remises=self.env['is.remise.particuliere'].search(filtre, limit=1)
            for remise in remises:
                if discount<remise.remise_client:
                    discount = remise.remise_client
            #******************************************************************
            obj.discount = discount


    @api.depends('product_id','product_uom_qty')
    def _compute_is_nb_pieces_par_colis(self):
        for obj in self:
            nb_colis = obj.get_nb_colis()
            obj.is_nb_colis = nb_colis
            unite = obj.product_uom.category_id.name
            if unite=="Poids":
                poids = obj.product_uom_qty
            else:
                poids = nb_colis*obj.product_id.is_poids_net_colis
            obj.is_poids_net = poids
            obj.is_nb_pieces_par_colis=obj.product_id.is_nb_pieces_par_colis
            if obj.discount==0:
                obj.get_discount()


    @api.depends('product_id','name','product_uom_qty')
    def _compute_ref(self):
        for obj in self:
            obj.is_default_code    = obj.product_id.default_code
            obj.is_ref_fournisseur = obj.product_id.is_ref_fournisseur


    @api.depends('product_id','name','product_uom_qty','qty_delivered')
    def _compute_is_colis_liv(self):
        for obj in self:
            colis_liv = 0
            filtre=[
                ('sale_line_id','=',obj.id),
                ('state'       ,'=','done'),
            ]
            moves = self.env['stock.move'].search(filtre)
            for move in moves:
                sens=1
                if move.picking_type_id.code=='incoming':
                    sens=-1
                colis_liv+= sens*move.is_nb_colis
            obj.is_colis_liv = colis_liv


    @api.depends('product_id','product_uom_qty','is_qt_cde')
    def _compute_is_ecart_qt_cde_prepa(self):
        for obj in self:
            ecart=obj.product_uom_qty - obj.is_qt_cde
            obj.is_ecart_qt_cde_prepa = ecart


    is_purchase_line_id       = fields.Many2one('purchase.order.line', string=u'Ligne commande fournisseur', index=True, copy=False)
    is_date_reception         = fields.Date(string=u'Date réception')
    is_livraison_directe      = fields.Boolean(string=u'Livraison directe', help=u"Si cette case est cochée, une commande fournisseur spécifique pour ce client sera créée",default=False)
    is_nb_pieces_par_colis    = fields.Integer(string="PCB", help='Nb Pièces / colis', compute='_compute_is_nb_pieces_par_colis', readonly=True, store=True)
    is_nb_colis               = fields.Float(string='Nb Colis', digits=(14,2), compute='_compute_is_nb_pieces_par_colis', readonly=True, store=True)
    is_colis_cde              = fields.Float(string='Colis Prépa', digits=(14,2))
    is_colis_cde_origine      = fields.Float(string='Colis Cde'  , digits=(14,2), readonly=True, help="Ce champ permet de mémoriser la valeur du champ 'Colis Prépa' au moment de la validation de la commande")
    is_poids_net              = fields.Float(string='Poids net', digits='Stock Weight', compute='_compute_is_nb_pieces_par_colis', readonly=True, store=True, help="Poids net total (Kg)")
    is_correction_prix_achat  = fields.Float(string="Correction Prix d'achat", digits='Product Price', help="Utilsé dans 'Lignes des mouvements valorisés'")
    is_default_code           = fields.Char(string='Réf Fromtome'   , compute='_compute_ref', readonly=True, store=True)
    is_ref_fournisseur        = fields.Char(string='Réf Fournisseur', compute='_compute_ref', readonly=True, store=True)
    is_colis_liv              = fields.Float(string='Colis Liv', digits=(14,2), compute='_compute_is_colis_liv', readonly=True, store=False)
    is_qt_cde                 = fields.Float(string='Qt Cde', digits='Product Unit of Measure',readonly=True,help="Ce champ permet de mémoriser la valeur du champ product_uom_qty au moment de la validation de la commande")
    is_ecart_qt_cde_prepa     = fields.Float(string='Qt Prépa - Qt Cde', digits='Product Unit of Measure', compute='_compute_is_ecart_qt_cde_prepa', readonly=True, store=True)


    def write(self, vals):
        res = super().write(vals)
        if 'product_uom_qty' in vals:
            for obj in self:
                filtre=[
                    ('sale_line_id','=',obj.id),
                ]
                moves = self.env['stock.move'].search(filtre)
                if len(moves)==1:
                    for move in moves:
                        if move.picking_id.state not in ['done','cancel']:
                            move.product_uom_qty = vals["product_uom_qty"]
        return res


    def _check_line_unlink(self):
        "Permet de désactiver ce controle pour pouvoir supprimer des lignes de commandes"
        res = self.filtered(lambda line: line.state in ('sale', 'done') and (line.invoice_lines or not line.is_downpayment) and not line.display_type)
        return False


    def unlink(self):
        for obj in self:
            moves=self.env['stock.move'].search([('sale_line_id', '=', obj.id)])
            for move in moves:
                if move.state not in ['cancel', 'done'] and move.picking_id.state not in ['cancel', 'done']:
                    move._action_cancel()
                    move.unlink()
                    msg = "Suppression de la ligne et du mouvement de stock associé pour %s " % (obj.product_id.display_name)
                    obj.order_id.message_post(body=msg)
                else:
                    raise UserError("Il n'est pas possible de supprimer une ligne livrée")
        res = super(SaleOrderLine, self).unlink()
        return res


    @api.onchange('product_uom_qty')
    def _onchange_product_uom_qty(self):
        # When modifying a one2many, _origin doesn't guarantee that its values will be the ones
        # in database. Hence, we need to explicitly read them from there.
        if self._origin:
            product_uom_qty_origin = self._origin.read(["product_uom_qty"])[0]["product_uom_qty"]
        else:
            product_uom_qty_origin = 0

        if self.state == 'sale' and self.product_id.type in ['product', 'consu'] and self.product_uom_qty < product_uom_qty_origin:
            # Do not display this warning if the new quantity is below the delivered
            # one; the `write` will raise an `UserError` anyway.
            if self.product_uom_qty < self.qty_delivered:
                return {}
            #warning_mess = {
            #    'title': _('Ordered quantity decreased!'),
            #    'message' : _('You are decreasing the ordered quantity! Do not forget to manually update the delivery order if needed.'),
            #}
            #return {'warning': warning_mess}
        if self.product_packaging:
            return self._check_package()
        return {}




    def get_fournisseur_par_defaut(self):
        now = datetime.date.today()
        suppliers=self.env['product.supplierinfo'].search([('product_tmpl_id', '=', self.product_id.product_tmpl_id.id)])
        partner=False
        for s in suppliers:
            if now>=s.date_start and now<= s.date_end:
                partner=s.name
                break
        return partner


    @api.onchange('is_colis_cde')
    def onchange_is_colis_cde(self):
        unite = self.product_uom.category_id.name
        if unite=="Poids":
            self.product_uom_qty =  self.is_colis_cde * self.product_id.is_poids_net_colis
        else:
            self.product_uom_qty =  self.is_colis_cde * self.is_nb_pieces_par_colis


    @api.onchange('product_uom_qty')
    def onchange_product_uom_qty_colis(self):
        if self.is_nb_pieces_par_colis>0:
            self.is_colis_cde = self.get_nb_colis()


    @api.onchange('product_id','is_livraison_directe')
    def onchange_product_id_for_date_reception(self):
        self.is_date_reception = datetime.date.today()+datetime.timedelta(days=2)

        #TODO : Désactivé le 08/08/2023 car cela pose problème pour Le Cellier
        # if self.is_livraison_directe:
        #     if self.order_id.is_date_livraison:
        #         self.is_date_reception = self.order_id.is_date_livraison
        # else:
        #     partner = self.get_fournisseur_par_defaut()
        #     if partner and partner.is_date_reception:
        #         self.is_date_reception = partner.is_date_reception


    # @api.onchange('is_date_reception')
    # def onchange_is_date_reception(self):
    #     if self.is_date_reception:
    #         partner = self.get_fournisseur_par_defaut()
    #         if partner:
    #             partner.write({'is_date_reception': self.is_date_reception})


    def acceder_commande_fournisseur(self):
        for obj in self:
            res= {
                'name': 'Ligne commande fournisseur',
                'view_mode': 'tree,form',
                'view_type': 'form',
                'res_model': 'purchase.order.line',
                'type': 'ir.actions.act_window',
                'domain': [
                    ('id','=',obj.is_purchase_line_id.id),
                ],
            }
            return res


class SaleOrder(models.Model):
    _inherit = "sale.order"


    def _compute_is_nb_lignes(self):
        for obj in self:
            nb=0
            for line in obj.order_line:
                if line.product_id!=obj.partner_id.is_frais_port_id:
                    if line.product_uom_qty>0:
                        nb+=1
            obj.is_nb_lignes = nb


    @api.depends('partner_id')
    def _compute_is_transporteur_id(self):
        for obj in self:
            obj.is_transporteur_id = obj.partner_id.is_transporteur_id.id
 

    is_enseigne_id           = fields.Many2one('is.enseigne.commerciale', 'Enseigne', related='partner_id.is_enseigne_id')
    is_date_livraison        = fields.Date('Date livraison client', help="Date d'arrivée chez le client", tracking=True)
    is_commande_soldee       = fields.Boolean(string='Commande soldée', default=False, copy=False, tracking=True, help="Cocher cette case pour indiquer qu'aucune nouvelle livraison n'est prévue sur celle-ci")
    is_frequence_facturation = fields.Selection(string='Fréquence facturation', related="partner_id.is_frequence_facturation") #, selection=[('au_mois', 'Au mois'),('a_la_livraison', 'A la livraison')])
    is_type_doc              = fields.Selection([('cc', 'CC'), ('offre', 'Offre')], string='Type document', default="cc", tracking=True)
    is_modele_commande_id    = fields.Many2one('is.modele.commande', 'Modèle de commande', related='partner_id.is_modele_commande_id', tracking=True)
    is_transporteur_id       = fields.Many2one('is.transporteur', 'Transporteur', compute='_compute_is_transporteur_id', store=True, readonly=False, tracking=True)
    is_encours_client        = fields.Float(related='partner_id.is_encours_client')
    is_import_excel_ids      = fields.Many2many('ir.attachment' , 'sale_order_is_import_excel_ids_rel', 'order_id'     , 'attachment_id'    , 'Commande .xlsx à importer')
    is_import_alerte         = fields.Text('Alertes importation')
    is_nb_lignes             = fields.Integer('Nb lignes (hors transport)', compute='_compute_is_nb_lignes')
    is_heure_envoi_id        = fields.Many2one('is.heure.maxi', 'Jour / Heure limite', tracking=True, help="Heure maxi d'envoi de la commande au fournisseur")
    is_fusion_order_id       = fields.Many2one('sale.order', 'Fusionnée dans', copy=False,readonly=True)
    is_frais_port_id           = fields.Many2one(related='partner_id.is_frais_port_id')


    def _message_auto_subscribe_notify(self, partner_ids, template):
        "Désactiver les notifications d'envoi des mails"
        return True


    def ajout_frais_de_port(self):
        "Ajout des frais de port"
        for order in self:
            is_mini_cde_franco = order.partner_id.is_mini_cde_franco or order.is_enseigne_id.mini_cde_franco
            if order.partner_id.is_frais_port_id:
                test=True
                for line in order.order_line:
                    if line.product_id==order.partner_id.is_frais_port_id:
                        test=False
                        port = order.partner_id.is_frais_port_id.lst_price
                        if is_mini_cde_franco>0 and (order.amount_untaxed - port)>=is_mini_cde_franco:
                            #** Pour supprimer la ligne, il ne faut pas de mouvement de stock lié
                            filtre=[
                                ('sale_line_id', '='     , line.id),
                            ]
                            moves=self.env['stock.move'].search(filtre)
                            if len(moves)>0:
                                for move in moves:
                                    if move.state not in ['cancel','done']:
                                        line.product_uom_qty = 0
                            else:
                                line.unlink()
                            #**************************************************
                        else:
                            line.product_uom_qty = 1
                        break
                if is_mini_cde_franco>0 and order.amount_untaxed>=is_mini_cde_franco:
                    test=False
                if test:
                    vals={
                        "order_id": order.id,
                        "sequence": 999,
                        "product_id": order.partner_id.is_frais_port_id.id,
                        "price_unit": order.partner_id.is_frais_port_id.lst_price,
                        "product_uom_qty": 1,
                    }
                    order_line = self.env['sale.order.line'].create(vals)
        return


    @api.model
    def create(self, vals):
        order = super(SaleOrder, self).create(vals)
        order.ajout_frais_de_port()
        return order


    def write(self, vals):
        res = super().write(vals)
        for obj in self:
            obj.ajout_frais_de_port()
            line_port = False
            for line in obj.order_line:
                if line.product_id==obj.partner_id.is_frais_port_id:
                    line_port=line
                    break
            if line_port:
                sequence=0
                for line in obj.order_line:
                    if line.sequence>sequence:
                        sequence = line.sequence
                sequence+=10
                line_port.sequence=sequence
            for picking in obj.picking_ids:
                picking.trier_par_emplacement_fournisseur()
        return res


    def action_confirm(self):
        for obj in self:
            obj.ajout_frais_de_port()
            for line in obj.order_line:
                if line.product_uom_qty==0:
                    line.unlink()
            obj.recopie_qt_prepa_dans_qt_cde_action(force=True)
            res = super(SaleOrder, self).action_confirm()
            for picking in obj.picking_ids:
                picking.trier_par_emplacement_fournisseur()
        return res


    @api.onchange('partner_id','company_id','user_id','is_enseigne_id')
    def onchange_partner_id_warehouse(self):
        warehouse_id=self.partner_id.is_warehouse_id.id or self.partner_id.is_enseigne_id.warehouse_id
        self.warehouse_id = warehouse_id

        # if self.partner_id and self.partner_id.is_warehouse_id:
        #     self.warehouse_id = self.partner_id.is_warehouse_id.id
        # else:
        #     if self.partner_id and self.partner_id.is_enseigne_id and self.partner_id.is_enseigne_id.warehouse_id:
        #         self.warehouse_id = self.partner_id.is_enseigne_id.warehouse_id.id
 

    @api.depends('order_line')
    def _compute_is_creer_commande_fournisseur_vsb(self):
        for obj in self:
            vsb = False
            for line in obj.order_line:
                if not line.is_purchase_line_id.id:
                    vsb=True
                    break
            obj.is_creer_commande_fournisseur_vsb=vsb


    is_creer_commande_fournisseur_vsb = fields.Boolean(string=u'Créer commande fournisseur', compute='_compute_is_creer_commande_fournisseur_vsb', readonly=True, store=False)


    def initialiser_depuis_modele_commande(self):
        date_reception = datetime.date.today()+datetime.timedelta(days=2)
        for obj in self:
            sequence=10
            for line in obj.is_modele_commande_id.ligne_ids:
                vals={
                    'order_id'    : obj.id,
                    'sequence'    : sequence,
                    'product_id'  : line.product_id.id,
                    'name'        : line.product_id.name_get()[0][1],
                    'product_uom_qty': 0,
                    'is_date_reception': date_reception,
                }
                self.env['sale.order.line'].create(vals)
                sequence+=10


    def commande_soldee_action_server(self):
        for obj in self:
            solde=False
            if obj.state=="sale":
                solde=True
                pickings = obj.picking_ids  
                for picking in pickings:
                    if picking.state not in ['done','cancel']:
                        solde=False
            obj.is_commande_soldee=solde


    def commande_entierement_facturee_action_server(self):
        for obj in self:
            # Forcer l'état de facturation au niveau des lignes de commande
            for line in obj.order_line:
                if not line.display_type:  # Exclure les lignes de section et note
                    line.invoice_status = "invoiced"
            # Le champ invoice_status de la commande sera automatiquement recalculé
            # grâce aux dépendances @api.depends


    # def _get_invoice_status(self):
    #     """
    #     Surcharge pour permettre le forçage manuel de l'état de facturation
    #     via commande_entierement_facturee_action_server
    #     """
    #     # Vérifier si toutes les lignes sont forcées en "invoiced"
    #     for order in self:
    #         all_lines_forced_invoiced = True
    #         for line in order.order_line:
    #             if not line.display_type and line.invoice_status != 'invoiced':
    #                 all_lines_forced_invoiced = False
    #                 break
            
    #         # Si toutes les lignes sont forcées en "invoiced", garder cet état
    #         if all_lines_forced_invoiced and order.state in ('sale', 'done'):
    #             order.invoice_status = 'invoiced'
    #             continue
        
    #     # Sinon, utiliser la logique standard d'Odoo
    #     super()._get_invoice_status()


    def initialisation_etat_facturee_action_server(self):
        for obj in self:
            for line in obj.order_line:
                line._compute_invoice_status()
 

    def recopie_qt_prepa_dans_qt_cde_action(self,force=False):
        for obj in self:
            for line in obj.order_line:
                if force:
                    line.is_qt_cde = line.product_uom_qty
                    line.is_colis_cde_origine = line.is_colis_cde
                else:
                    if line.is_qt_cde==0:
                        line.is_qt_cde = line.product_uom_qty
                    if line.is_colis_cde_origine==0:
                        line.is_colis_cde_origine = line.is_colis_cde

          
    def creer_commande_fournisseur_action(self):
        company = self.env.user.company_id
        for obj in self:
            if not len(obj.order_line):
                raise Warning(u"Il n'y a aucune ligne de commandes à traiter !")
            for line in obj.order_line:
                if not line.is_date_reception:
                    raise Warning(u"La date de réception n'est pas renseignée sur toutes les lignes")
            now = datetime.date.today()
            for line in obj.order_line:
                _logger.info("creer_commande_fournisseur_action : article=%s (%s)"%(line.product_id.display_name,line.product_id.id))
                if not line.is_purchase_line_id:
                    suppliers=self.env['product.supplierinfo'].search([('product_tmpl_id', '=', line.product_id.product_tmpl_id.id)])
                    partner_id=False
                    supplierinfo=False
                    for s in suppliers:
                        if now>=s.date_start and now<= s.date_end:
                            supplierinfo=s
                            break
                    if not supplierinfo:
                        raise Warning("Fournisseur non trouvé pour article '%s'"%(line.product_id.display_name))
                    if supplierinfo:
                        if not supplierinfo.name.active:
                            raise Warning("Fournisseur '%s' désactivé pour l'article '%s'"%(supplierinfo.name.name,line.product_id.display_name))
                        if not supplierinfo.name.is_warehouse_id:
                            raise Warning("Entrepôt non renseigné pour le fournisseur '%s' de l'article '%s'"%(supplierinfo.name.name,line.product_id.display_name))
                        if supplierinfo.name.is_heure_envoi_id==obj.is_heure_envoi_id or obj.is_heure_envoi_id.id==False:
                            partner_id = supplierinfo.name.id
                            date_reception = str(line.is_date_reception)
                            date_planned  = date_reception+' 08:00:00'
                            filtre=[
                                ('partner_id'  ,'='   , partner_id),
                                ('state'       ,'='   , 'draft'),
                                ('date_planned','>=', date_reception+' 00:00:00'),
                                ('date_planned','<=', date_reception+' 23:59:59'),
                            ]
                            if line.is_livraison_directe:
                                filtre.append(('is_adresse_livraison_id','=', obj.partner_shipping_id.id))
                            orders=self.env['purchase.order'].search(filtre,limit=1)
                            if orders:
                                order=orders[0]
                            else:
                                vals={
                                    'partner_id'             : partner_id,
                                    'date_planned'           : date_planned,
                                    'picking_type_id'        : supplierinfo.name.is_warehouse_id.in_type_id.id,
                                    'is_adresse_livraison_id': supplierinfo.name.is_enseigne_id.warehouse_id.partner_id.id,  #supplierinfo.name.is_enseigne_id.name.id,
                                }
                                order=self.env['purchase.order'].create(vals)
                                if order:
                                    order.onchange_partner_id()
                                    if line.is_livraison_directe:
                                        order.is_adresse_livraison_id = obj.partner_shipping_id.id

                            #** Création des lignes ************************************
                            if company.is_regroupe_cde=="Oui":
                                filtre=[
                                    ('order_id'  ,'=', order.id),
                                    ('product_id','=', line.product_id.id),
                                ]
                            else:
                                filtre=[
                                    ('order_id'             ,'=' , order.id),
                                    ('product_id'           ,'=' , line.product_id.id),
                                    ('is_sale_order_line_id','=' , line.id),
                                    ('is_sale_order_line_id','!=', False),
                                ]
                            order_lines=self.env['purchase.order.line'].search(filtre,limit=1)
                            if not order_lines:
                                if order:
                                    vals={
                                        'order_id'    : order.id,
                                        'product_id'  : line.product_id.id,
                                        'name'        : line.name,
                                        'product_qty' : line.product_uom_qty,
                                        'product_uom' : line.product_uom.id,
                                        'date_planned': date_planned,
                                        'price_unit'  : line.price_unit,
                                    }

                                    if company.is_regroupe_cde=="Non":
                                        vals["is_sale_order_line_id"]=line.id

                                    order_line=self.env['purchase.order.line'].create(vals)
                                    order_line.onchange_product_id()
                                    order_line.onchange_product_qty_fromtome()
                                    order_line.date_planned = date_planned
                            else:
                                order_line = order_lines[0]
                            if order_line:
                                line.is_purchase_line_id=order_line.id
                                if company.is_regroupe_cde=="Non":
                                    order_line.product_qty = line.product_uom_qty
                                else:
                                    filtre=[
                                        ('is_purchase_line_id'  ,'=', order_line.id),
                                    ]
                                    order_lines=self.env['sale.order.line'].search(filtre)
                                    qty=0
                                    for l in order_lines:
                                        qty+=l.product_uom_qty
                                    order_line.product_qty = qty
                                    order_line.onchange_product_qty_fromtome()
                            # ***********************************************************


    def import_fichier_xlsx(self):
        for obj in self:
            #obj.order_line.unlink()
            alertes=[]
            for attachment in obj.is_import_excel_ids:
                xlsxfile=base64.b64decode(attachment.datas)

                path = '/tmp/sale_order-'+str(obj.id)+'.xlsx'
                f = open(path,'wb')
                f.write(xlsxfile)
                f.close()
                #*******************************************************************

                #** Test si fichier est bien du xlsx *******************************
                try:
                    wb    = load_workbook(filename = path, data_only=True)
                    ws    = wb.active
                    cells = list(ws)
                except:
                    raise Warning(u"Le fichier "+attachment.name+u" n'est pas un fichier xlsx")
                #*******************************************************************

                lig=0
                for row in ws.rows:
                    if lig>5:
                        num         = cells[lig][0].value
                        designation = cells[lig][1].value
                        code        = cells[lig][2].value
                        colis       = cells[lig][6].value
                        try:
                            colis = float(colis or 0)
                        except ValueError:
                            colis = 0
                        if colis and colis>0:
                            if code==None or code==False:
                                alertes.append("Ligne %s : Code article non indiqué sur la ligne"%(lig+1))
                            else:
                                filtre=[
                                    ("default_code","=", code),
                                ]
                                products = self.env['product.product'].search(filtre, limit=1)
                                if len(products)>0:
                                    product=products[0]
                                    try:
                                        qty = colis
                                    except ValueError:
                                        qty = 0
                                    vals={
                                        "order_id"         : obj.id,
                                        "product_id"       : product.id,
                                        "sequence"         : lig,
                                        "is_colis_cde"     : qty,
                                        "product_uom"      : product.uom_id.id,
                                        "is_date_reception": datetime.date.today()+datetime.timedelta(days=2),
                                    }
                                    res = self.env['sale.order.line'].create(vals)
                                    res.onchange_is_colis_cde()
                                else:
                                    alertes.append("Ligne %s : Article %s non trouvé"%((lig+1),code))
                    lig+=1
            obj.ajout_frais_de_port()
            if alertes:
                alertes = "\n".join(alertes)
            else:
                alertes=False
            obj.is_import_alerte = alertes





    def trier_par_designation_action(self):
        for obj in self:
            my_dict={}
            for move in obj.order_line:
                if move.product_id.default_code:
                    name=move.product_id.name
                else:
                    name="zzzz"
                key="%s-%s"%(name, move.id)
                my_dict[key]=move
            sorted_dict = dict(sorted(my_dict.items()))
            sequence=10
            for key in sorted_dict:
                move=sorted_dict[key]
                move.sequence=sequence
                sequence+=10


    def trier_par_ref_fromtome_action(self):
        for obj in self:
            my_dict={}
            for move in obj.order_line:
                key="%s-%s"%(move.product_id.default_code or 'zzzz', move.id)
                my_dict[key]=move
            sorted_dict = dict(sorted(my_dict.items()))
            sequence=10
            for key in sorted_dict:
                move=sorted_dict[key]
                move.sequence=sequence
                sequence+=10


    def trier_par_ref_fournisseur_action(self):
        for obj in self:
            my_dict={}
            for move in obj.order_line:
                key="%s-%s"%(move.is_ref_fournisseur or 'zzzz', move.id)
                my_dict[key]=move
            sorted_dict = dict(sorted(my_dict.items()))
            sequence=10
            for key in sorted_dict:
                move=sorted_dict[key]
                move.sequence=sequence
                sequence+=10


    def trier_par_poids_action(self):
        for obj in self:
            my_dict={}
            for move in obj.order_line:
                poids=str(int(move.product_id.is_poids_net_colis*100000)).zfill(10)
                key="%s-%s"%(poids, move.id)
                my_dict[key]=move
            sorted_dict = dict(sorted(my_dict.items(),reverse=True))
            sequence=10
            for key in sorted_dict:
                move=sorted_dict[key]
                move.sequence=sequence
                sequence+=10






    def fusion_commande_client_action(self):
        "Permet de fusionner les commandes validées et non livrées de la même date et du même client"
        dict={}
        #** Recherche des commandes par client et par date **************
        for obj in self:
            if obj.partner_id and obj.is_date_livraison and obj.state!='cancel':
                test=False
                if obj.state in ['draft', 'sent']:
                    test=True
                else:
                    for picking in obj.picking_ids:
                        if picking.state in ['done','cancel']:
                            test=True
                            break
                if test:
                    key = "%s-%s"%(obj.partner_id.id, obj.is_date_livraison.strftime('%Y-%m-%d'))
                    if key not in dict:
                        dict[key]=[]
                    dict[key].append(obj.id)
        #*** Suppresion des clés avec une seule commande => Pas de fusion *****
        for key in list(dict):
            if len(dict[key])==1:
                dict.pop(key)

        #** Fusion des commandes **********************************************
        for key in dict:
            first_order=False
            orders = dict[key]
            #orders.sort() #Cela permet de conserver la première commande et d'annuler les autres
            for order_id in orders:
                order = self.env['sale.order'].browse(order_id)
                if not first_order:
                    first_order=order
                else:
                    #** Recherche sequence pour mettre la ligne à la fin ******
                    sequence=0
                    for line in first_order.order_line:
                        if line.sequence>sequence:
                            sequence=line.sequence
                    #**********************************************************
                    for line in order.order_line:

                        #** Recherche si ligne existe pour cet article ********
                        test=False
                        for l in first_order.order_line:
                            if l.product_id == line.product_id:
                                qty = line.product_uom_qty + l.product_uom_qty
                                l.product_uom_qty = qty
                                l.onchange_product_uom_qty_colis()
                                test=True
                                break
                        #** Création d'une nouvelle ligne *********************
                        if test==False:
                            sequence+=10
                            vals={
                                'order_id'    : first_order.id,
                                'sequence'    : sequence,
                                'product_id'  : line.product_id.id,
                                'name'        : line.name,
                                'product_uom_qty' : line.product_uom_qty,
                                'product_uom' : line.product_uom.id,
                            }
                            order_line=self.env['sale.order.line'].create(vals)
                            order_line.onchange_product_uom_qty_colis()

                    order.action_cancel()
                    first_order.trier_par_designation_action()
                    order.is_fusion_order_id = first_order.id
                    msg = "Commande annulée et fusionnée avec %s" % (first_order.name)
                    order.message_post(body=msg)
                    msg = "Fusion de la commande %s" % (order.name)
                    first_order.message_post(body=msg)

