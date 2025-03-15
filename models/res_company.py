from odoo import api, fields, models  # type: ignore
from email.policy import default
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Fill, Alignment,PatternFill
from datetime import datetime
import pytz
import base64
import logging
_logger = logging.getLogger(__name__)


class Company(models.Model):
    _inherit = 'res.company'

    is_gln          = fields.Char(string='GLN')
    is_regroupe_cde = fields.Selection(string='Regrouper les commandes', selection=[('Oui', 'Oui'), ('Non', 'Non')], default="Oui", help="Regrouper les commandes des clients sur les commandes fournisseur")

    is_date_bascule_tarif = fields.Date(string='Date bascule tarif client')

    is_coef_cdf_quai      = fields.Float(string='Taux de marge A QUAI (%)'     , digits=(14,2))
    is_coef_cdf_franco    = fields.Float(string='Taux de marge FRANCO (%)'   , digits=(14,2))
    is_coef_ft            = fields.Float(string='Taux de marge FROMTOME  (%)', digits=(14,2))
    is_coef_lf            = fields.Float(string='Taux de marge LF (%)        (Ne plus utiliser)', digits=(14,2))
    is_coef_lf_coll       = fields.Float(string='Taux de marge LF coll. (%)  (Ne plus utiliser)', digits=(14,2))
    is_coef_lf_franco     = fields.Float(string='Taux de marge LF franco (%) (Ne plus utiliser)', digits=(14,2))


    is_port_cdf_quai      = fields.Float(string='Frais de port A QUAI (€)'  , digits=(14,2))
    is_port_cdf_franco    = fields.Float(string='Frais de port FRANCO (€)'  , digits=(14,2))
    is_port_ft            = fields.Float(string='Frais de port FROMTOME (€)', digits=(14,2))
    is_port_lf            = fields.Float(string='Frais de port LF (€)        (Ne plus utiliser)', digits=(14,2))
    is_port_lf_coll       = fields.Float(string='Frais de port LF coll. (€)  (Ne plus utiliser)', digits=(14,2))
    is_port_lf_franco     = fields.Float(string='Frais de port LF franco (€) (Ne plus utiliser)', digits=(14,2))

    is_import_excel_ids   = fields.Many2many('ir.attachment' , 'res_company_is_import_excel_ids_rel', 'company_id', 'attachment_id'    , 'Tarifs .xlsx à importer')
    is_import_alerte      = fields.Text('Alertes importation')
    is_mini_cde_franco    = fields.Integer(string='Mini de commande franco (€)')


    def actualiser_tarif_futur_action(self):
        for obj in self:
            _logger.info("actualiser_tarif_futur_action : Début")
            self.env['product.template'].search([])._compute_tarifs(update_prix_actuel=False)
            #self.send_mail('actualiser_tarif_futur_action')
            _logger.info("actualiser_tarif_futur_action : Fin")


    def appliquer_nouveaux_tarifs_action(self):
        self.env['product.template'].appliquer_nouveaux_tarifs_action()
        #self.send_mail('appliquer_nouveaux_tarifs_action')


    # def actualiser_tarif_action(self):
    #     for obj in self:
    #         self.env['product.template'].search([])._compute_tarifs(update_prix_actuel=True)
    #         self.send_mail('actualiser_tarif_action')



    # def actualiser_tarif_ft_action(self):
    #     for obj in self:
    #         self.env['product.template'].search([])._compute_tarifs(update_prix_actuel=True, pricelist='ft')
    #         #self.send_mail('actualiser_tarif_action')


    # def actualiser_tarif_lf_franco_action(self):
    #     for obj in self:
    #         self.env['product.template'].search([])._compute_tarifs(update_prix_actuel=True, pricelist='lf_franco')
    #         #self.send_mail('actualiser_tarif_action')


    def send_mail(self,action): 
        tz = pytz.timezone('Europe/Paris')
        now = datetime.now(tz).strftime("%d/%m/%Y à %H:%M:%S")
        subject = "Action %s terminée le %s"%(action,now)
        body="""
            <p>Bonjour,</p> 
            <p>Traitement terminé</p> 
        """
        ctx = dict(
            default_model="res.company",
            default_res_id=self.id,
            default_composition_mode='comment',
            custom_layout='mail.mail_notification_light', #Permet de définir la mise en page du mail
        )
        infosaone_id=3128
        vals={
            "model"         : "res.company",
            "subject"       : subject,
            "body"          : body,
            "partner_ids"   : [self.env.user.partner_id.id,infosaone_id],
        }
        wizard = self.env['mail.compose.message'].with_context(ctx).create(vals)
        wizard.send_mail()
        return True




    def import_fichier_xlsx(self):
        for obj in self:
            alertes=[]
            product_ids=[]
            for attachment in obj.is_import_excel_ids:
                xlsxfile=base64.b64decode(attachment.datas)

                path = '/tmp/sale_order-'+str(obj.id)+'.xlsx'
                f = open(path,'wb')
                f.write(xlsxfile)
                f.close()
                #*******************************************************************

                #** Test si fichier est bien du xlsx *******************************
                try:
                    wb    = load_workbook(filename = path, data_only=True)
                    ws    = wb.active
                    cells = list(ws)
                    row_count = ws.max_row
                except:
                    raise Warning(u"Le fichier "+attachment.name+u" n'est pas un fichier xlsx")
                #*******************************************************************

                lig=0
                nb=row_count
                for row in ws.rows:
                    if lig>0:
                        default_code = cells[lig][0].value
                        filtre=[
                            ("default_code","=", default_code),
                        ]
                        products = self.env['product.template'].search(filtre, limit=1)
                        if len(products)==0:
                            msg="%s : Référence non trouvée"%default_code
                            alertes.append(msg)
                        else:
                            taux_marge  = cells[lig][1].value
                            try:
                                taux_marge = round(100*float(taux_marge or 0),4)
                            except ValueError:
                                taux_marge = 0
                            if taux_marge==0:
                                msg="%s : Taux de marge à 0"%default_code
                                alertes.append(msg)
                            if taux_marge<5:
                                msg="%s : Taux de marge <5 (%s)"%(default_code,taux_marge)
                                alertes.append(msg)
                            if taux_marge>50:
                                msg="%s : Taux de marge >50 (%s)"%(default_code,taux_marge)
                                alertes.append(msg)
                            if taux_marge>0:
                                product=products[0]
                                vals={
                                    'is_prix_vente_futur_marge_cdf_quai'  : taux_marge,
                                    'is_prix_vente_futur_marge_cdf_franco': taux_marge,
                                    'is_prix_vente_futur_marge_ft'        : taux_marge,
                                }
                                product.write(vals)
                                product_ids.append(product)
                                _logger.info("import_fichier_xlsx : %s/%s : %s : %s"%(lig,nb,default_code,taux_marge))
                    lig+=1
                    #if lig>50:
                    #    break
            if alertes:
                for alerte in alertes:
                    _logger.info("import_fichier_xlsx : Alerte : %s"%alerte)
                alertes = "\n".join(alertes)
            else:
                alertes=False
            obj.is_import_alerte = alertes
            if len(product_ids)>0:
                for product in product_ids:
                    product._compute_tarifs(update_prix_actuel=False)


